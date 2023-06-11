import serial
import time
import tkinter as tk
import serial.tools.list_ports
import threading
from tkinter import filedialog
from openpyxl import Workbook, load_workbook
import os.path
import matplotlib.pyplot as plt
from PIL import Image, ImageTk


class MotorController:
    def __init__(self):
        self.ser = None
        self.com_var = None
        self.distance_var = None
        self.distance_label = None
        self.distance_entry = None
        self.start_button = None
        self.stop_button = None
        self.halt_var = None
        self.halt_label = None
        self.halt_entry = None
        self.ports = get_serial_ports()

        self.root = tk.Tk()
        self.root.geometry('1600x300')
        self.root.title("Motor Controller")
        self.root.iconbitmap("resources/icon-removebg.ico")
        self.logo = Image.open("resources/image-removebg.png.png")

        #TODO: Please manually change the com port to Arduino COM port
        # we have set it to COM5, as default

        self.serial = serial.Serial('COM5', 9600, timeout=3)
        time.sleep(2)
        with open("./arduino/sketch_mar14a.ino", 'r') as f:
          code = f.read()
          self.serial.write(code.encode())
        self.filename = 'weight_data.xlsx'

        # check if the file already exists
        if os.path.isfile(self.filename):
            self.wb = load_workbook(filename=self.filename)
            self.ws = self.wb.active
        else:
           self.wb = Workbook()
           self.ws = self.wb.active
           self.ws.append(['Time', 'Weight (g)','Result'])
           self.wb.save(self.filename)
        #
        # Create dropdown menu for selecting serial port
        self.logo = self.logo.resize((150, 150))  # resize the image
        self.logo_tk = ImageTk.PhotoImage(self.logo)
        self.logo_label = tk.Label(self.root, image=self.logo_tk)
        self.logo_label.pack(side=tk.TOP, padx=10, pady=10, anchor='n')
        tk.Label(self.root, text="Select COM port:").pack(side=tk.TOP , anchor='nw')

        self.com_var = tk.StringVar()
        self.com_var.set(self.ports[0])
        com_dropdown = tk.OptionMenu(self.root, self.com_var, *self.ports)
        com_dropdown.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        # Add connect button
        tk.Button(self.root, text="Connect", command=self.connect).pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        # Add motor control UI elements here
        self.distance_var = tk.StringVar()
        self.distance_label = tk.Label(text="Displacement (cm):")
        self.distance_label.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')
        self.distance_entry = tk.Entry(textvariable=self.distance_var)
        self.distance_entry.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        self.halt_var = tk.StringVar()
        self.halt_label = tk.Label(text="Halt time (s):")
        self.halt_label.pack(side=tk.LEFT, anchor='nw',padx=10, pady=10)
        self.halt_entry = tk.Entry(textvariable=self.halt_var)
        self.halt_entry.pack(side=tk.LEFT,  anchor='nw', padx=10, pady=10)
        self.halt_var.set('5')

        self.weight_label = tk.Label(text='Load (grams-F) : ')
        self.weight_label.pack(side=tk.BOTTOM ,anchor='nw', padx=10, pady=10)

        self.max_weight_var = tk.StringVar()
        self.max_weight_label = tk.Label(text="Maximum load in grams-F:")
        self.max_weight_label.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')
        self.max_weight_entry = tk.Entry(textvariable=self.max_weight_var)
        self.max_weight_entry.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        self.min_weight_var = tk.StringVar()
        self.min_weight_label = tk.Label(text="Minimum load in grams-F:")
        self.min_weight_label.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')
        self.min_weight_entry = tk.Entry(textvariable=self.min_weight_var)
        self.min_weight_entry.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        self.start_button = tk.Button(text="Start", command=self.start_motor)
        self.start_button.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        self.stop_button = tk.Button(text="Stop", command=self.stop_motor)
        self.stop_button.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        self.graph_button = tk.Button(text="Generate Graph", command=self.generate_graph)
        self.graph_button.pack(side=tk.LEFT, padx=10, pady=10, anchor='nw')

        


    def connect(self):
        if self.ser is None:
            # Connect to selected serial port
            self.ser = serial.Serial(self.com_var.get(), 115200, timeout=1)
            print(f"Connected to {self.ser.name}")

    def start_motor(self):
        if self.ser is not None:
            distance_cm = int(self.distance_var.get())
            steps = distance_cm * 700
            halt_time = float(self.halt_var.get())
            max_weight = float(self.max_weight_var.get())
            min_weight = float(self.min_weight_var.get())

            self.motor_thread_stop = threading.Event()

            def motor_thread():
                while not self.motor_thread_stop.is_set():
                    # Move the motor down and halt for the specified time
                    self.send_command(str(steps))
                    time.sleep(halt_time)
                    self.save_data()
                    
                    # Move the motor up to its original position and halt for the specified time
                    self.send_command(str(-steps))
                    time.sleep(halt_time)

            self.motor_thread = threading.Thread(target=motor_thread)
            self.motor_thread.daemon = True
            self.motor_thread.start()
            self.read_weight(max_weight, min_weight)
            

    def stop_motor(self):
        if hasattr(self, 'motor_thread_stop'):
            self.motor_thread_stop.set()
        if hasattr(self, 'motor_thread'):
            self.motor_thread.join()
        if self.ser is not None:
            self.ser.close()
            self.ser = None
        self.motor_running = False

    def send_command(self, cmd):
        if self.ser is not None:
            self.ser.write(cmd.encode())
            time.sleep(0.1)
            response = self.ser.readline().decode().strip()
            return response
            

    def read_weight(self, max_weight, min_weight):
        line = self.serial.readline().decode().strip()
        if self.weight_label.winfo_exists() and line.startswith('Weight:'):
            weight = float(line.split()[1])
            self.weight_label.configure(text='Weight: {:.2f} g'.format(weight))
            if weight >= min_weight and weight <= max_weight:
                self.weight_label.configure(text='Weight: {:.2f} g - OK'.format(weight), fg='green')
            else:
                self.weight_label.configure(text='Weight: {:.2f} g - NOT OK'.format(weight), fg='red')

            # update data to be saved
            self.current_time = time.strftime('%Y-%m-%d %H:%M:%S')
            self.current_weight = weight
        self.root.after(100, self.read_weight, max_weight, min_weight)


    
    def save_data(self):
        weight = float(self.weight_label["text"].split()[1])
        result = "ok" if float(self.min_weight_var.get()) <= weight <= float(self.max_weight_var.get()) else "not ok"
        row = [time.strftime("%Y-%m-%d %H:%M:%S"), weight, result]
        self.ws.append(row)
        self.wb.save(self.filename)
    

    def generate_graph(self):
        weights = []
        times = []
        for row in self.ws.iter_rows(min_row=2, values_only=True):
            times.append(row[0])
            weights.append(row[1])
        
        plt.plot(times, weights)
        plt.xlabel('Time')
        plt.ylabel('Load (grams-F)')
        plt.title('Load vs Time')
        plt.show()
    
    def reset_graph(self):
        # Clear the plot
        plt.clf()
        
        # Reset the plot data
        self.x_data = []
        self.y_data = []
        self.ax = plt.axes()
        self.ax.set_xlabel('Time (s)')
        self.ax.set_ylabel('Load (grams-F)')
        self.ax.set_ylim(0, 300)
        self.line, = self.ax.plot(self.x_data, self.y_data, 'b-')
    
    # Redraw the plot
    plt.draw()
def get_serial_ports():
    """Returns a list of available serial ports."""
    ports = serial.tools.list_ports.comports()
    return [port.device for port in ports]


